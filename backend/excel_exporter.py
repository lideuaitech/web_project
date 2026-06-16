from openpyxl import Workbook
from openpyxl.styles import Font
import uuid
import os


class ExcelExporter:

    def __init__(self, df, summary):

        self.df = df
        self.summary = summary

    def export(self):

        wb = Workbook()

        ws = wb.active

        ws.title = "Analytics Report"

        # =========================
        # TITLE
        # =========================

        ws["A1"] = "Lideu AI Analytics Report"

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        # =========================
        # TABLE HEADERS
        # =========================

        headers = list(self.df.columns)

        row_start = 3

        for col_num, header in enumerate(headers, 1):

            cell = ws.cell(
                row=row_start,
                column=col_num
            )

            cell.value = header

            cell.font = Font(bold=True)

        # =========================
        # TABLE DATA
        # =========================

        for row_idx, row in enumerate(
            self.df.values,
            row_start + 1
        ):

            for col_idx, value in enumerate(
                row,
                1
            ):

                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=value
                )

        # =========================
        # SUMMARY
        # =========================

        summary_row = len(self.df) + 6

        ws.cell(
            row=summary_row,
            column=1,
            value="AI Insight Summary"
        ).font = Font(
            bold=True,
            size=14
        )

        ws.cell(
            row=summary_row + 1,
            column=1,
            value=self.summary
        )

        # =========================
        # AUTO WIDTH
        # =========================

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 5

        # =========================
        # SAVE FILE
        # =========================

        if not os.path.exists("exports"):
            os.makedirs("exports")

        filename = (
            f"exports/{uuid.uuid4()}.xlsx"
        )

        wb.save(filename)

        return filename